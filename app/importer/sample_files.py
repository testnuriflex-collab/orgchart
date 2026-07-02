from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SAMPLE_ROWS = [
    {
        "사번": "CEO001",
        "이름": "정민서",
        "이메일": "minseo.jung@example.com",
        "부서": "대표이사실",
        "상위부서": "",
        "직급": "대표",
        "직책": "CEO",
        "입사일": "2020-03-02",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "",
        "스킬": "전략, 투자, 리더십",
        "비고": "최상위 조직장",
    },
    {
        "사번": "HR001",
        "이름": "박서연",
        "이메일": "seoyeon.park@example.com",
        "부서": "People팀",
        "상위부서": "경영지원본부",
        "직급": "책임",
        "직책": "팀장",
        "입사일": "2021-01-11",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "FIN001",
        "스킬": "채용, 평가, 노무",
        "비고": "조직 개편 담당",
    },
    {
        "사번": "HR002",
        "이름": "한지우",
        "이메일": "jiwoo.han@example.com",
        "부서": "People팀",
        "상위부서": "경영지원본부",
        "직급": "매니저",
        "직책": "HR Operations",
        "입사일": "2023-04-03",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "HR001",
        "스킬": "입퇴사, 온보딩, 문서관리",
        "비고": "",
    },
    {
        "사번": "FIN001",
        "이름": "최도윤",
        "이메일": "doyun.choi@example.com",
        "부서": "재무회계팀",
        "상위부서": "경영지원본부",
        "직급": "이사",
        "직책": "경영지원본부장",
        "입사일": "2020-10-12",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "CEO001",
        "스킬": "재무, 예산, 내부통제",
        "비고": "본부장 겸 재무 리드",
    },
    {
        "사번": "ENG001",
        "이름": "김하준",
        "이메일": "hajun.kim@example.com",
        "부서": "플랫폼개발팀",
        "상위부서": "제품본부",
        "직급": "수석",
        "직책": "팀장",
        "입사일": "2021-05-17",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "PROD001",
        "스킬": "아키텍처, Python, 데이터",
        "비고": "대형 팀 리더",
    },
    {
        "사번": "ENG002",
        "이름": "이서준",
        "이메일": "seojun.lee@example.com",
        "부서": "플랫폼개발팀",
        "상위부서": "제품본부",
        "직급": "책임",
        "직책": "Backend Engineer",
        "입사일": "2022-09-05",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "ENG001",
        "스킬": "API, SQLite, 배포",
        "비고": "",
    },
    {
        "사번": "ENG003",
        "이름": "오유나",
        "이메일": "yuna.oh@example.com",
        "부서": "프론트엔드팀",
        "상위부서": "제품본부",
        "직급": "선임",
        "직책": "Frontend Engineer",
        "입사일": "2023-02-13",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "부산",
        "고용형태": "정규직",
        "매니저사번": "PROD001",
        "스킬": "Qt, UX, 접근성",
        "비고": "원격 근무",
    },
    {
        "사번": "ENG004",
        "이름": "김민준",
        "이메일": "minjun.kim.eng@example.com",
        "부서": "AI응용팀",
        "상위부서": "제품본부",
        "직급": "선임",
        "직책": "AI Engineer",
        "입사일": "2024-08-19",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "PROD001",
        "스킬": "LLM, 자동화, 평가",
        "비고": "동명이인 검증용",
    },
    {
        "사번": "PROD001",
        "이름": "강지훈",
        "이메일": "jihoon.kang@example.com",
        "부서": "프로덕트전략팀",
        "상위부서": "제품본부",
        "직급": "상무",
        "직책": "제품본부장",
        "입사일": "2020-08-24",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "CEO001",
        "스킬": "로드맵, 지표, 조직설계",
        "비고": "",
    },
    {
        "사번": "DSN001",
        "이름": "윤아린",
        "이메일": "arin.yoon@example.com",
        "부서": "프로덕트디자인팀",
        "상위부서": "제품본부",
        "직급": "선임",
        "직책": "Product Designer",
        "입사일": "2023-11-06",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "PROD001",
        "스킬": "디자인시스템, 리서치, 프로토타이핑",
        "비고": "Airbnb 톤 검증용",
    },
    {
        "사번": "SALES001",
        "이름": "문태오",
        "이메일": "taeo.moon@example.com",
        "부서": "세일즈팀",
        "상위부서": "사업본부",
        "직급": "책임",
        "직책": "팀장",
        "입사일": "2022-04-01",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "BIZ001",
        "스킬": "B2B 영업, 파이프라인, 협상",
        "비고": "",
    },
    {
        "사번": "BIZ001",
        "이름": "김민준",
        "이메일": "minjun.kim.biz@example.com",
        "부서": "사업전략팀",
        "상위부서": "사업본부",
        "직급": "이사",
        "직책": "사업본부장",
        "입사일": "2021-07-19",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "CEO001",
        "스킬": "사업개발, 파트너십, 매출관리",
        "비고": "동명이인 검증용",
    },
    {
        "사번": "CS001",
        "이름": "배수빈",
        "이메일": "subin.bae@example.com",
        "부서": "고객성공팀",
        "상위부서": "사업본부",
        "직급": "매니저",
        "직책": "Customer Success Manager",
        "입사일": "2024-03-04",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "부산",
        "고용형태": "정규직",
        "매니저사번": "BIZ001",
        "스킬": "온보딩, VOC, 리텐션",
        "비고": "",
    },
    {
        "사번": "MKT001",
        "이름": "임나은",
        "이메일": "naeun.lim@example.com",
        "부서": "브랜드마케팅팀",
        "상위부서": "사업본부",
        "직급": "선임",
        "직책": "Brand Marketer",
        "입사일": "2022-01-03",
        "퇴사일": "",
        "재직상태": "재직",
        "근무지": "서울",
        "고용형태": "계약직",
        "매니저사번": "BIZ001",
        "스킬": "캠페인, 콘텐츠, 브랜딩",
        "비고": "고용형태 필터 검증용",
    },
    {
        "사번": "OPS001",
        "이름": "장현우",
        "이메일": "hyunwoo.jang@example.com",
        "부서": "운영기획팀",
        "상위부서": "사업본부",
        "직급": "매니저",
        "직책": "Business Operations",
        "입사일": "2021-12-06",
        "퇴사일": "2026-05-31",
        "재직상태": "퇴사",
        "근무지": "서울",
        "고용형태": "정규직",
        "매니저사번": "BIZ001",
        "스킬": "프로세스, 정책, 리포팅",
        "비고": "퇴사자 표시 검증용",
    },
]


def write_sample_input_files(target_dir: Path) -> list[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame(SAMPLE_ROWS)
    paths = [
        target_dir / "인사정보_샘플.xlsx",
        target_dir / "인사정보_샘플.csv",
        target_dir / "인사정보_샘플.json",
    ]
    frame.to_excel(paths[0], index=False)
    frame.to_csv(paths[1], index=False, encoding="utf-8-sig")
    with paths[2].open("w", encoding="utf-8") as file:
        json.dump({"employees": SAMPLE_ROWS}, file, ensure_ascii=False, indent=2)
    return paths


# ── 표준 조직도 템플릿(2 시트: 명단 + 위계) ────────────────────────────
TEMPLATE_COMPANY = "(주)오르그스튜디오"

TEMPLATE_PEOPLE_ROWS = [
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "", "소속부서": "대표이사실",
     "이름": "정민서", "직책": "CEO", "직급": "대표", "사번": "CEO001",
     "이메일": "minseo.jung@example.com", "재직상태": "재직", "입사일": "2020-03-02", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "경영지원본부", "소속부서": "People팀",
     "이름": "박서연", "직책": "팀장", "직급": "책임", "사번": "HR001",
     "이메일": "seoyeon.park@example.com", "재직상태": "재직", "입사일": "2021-01-11", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "경영지원본부", "소속부서": "People팀",
     "이름": "한지우", "직책": "HR Operations", "직급": "매니저", "사번": "HR002",
     "이메일": "jiwoo.han@example.com", "재직상태": "재직", "입사일": "2023-04-03", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "경영지원본부", "소속부서": "재무회계팀",
     "이름": "최도윤", "직책": "본부장", "직급": "이사", "사번": "FIN001",
     "이메일": "doyun.choi@example.com", "재직상태": "재직", "입사일": "2020-10-12", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "제품본부", "소속부서": "플랫폼개발팀",
     "이름": "김하준", "직책": "팀장", "직급": "수석", "사번": "ENG001",
     "이메일": "hajun.kim@example.com", "재직상태": "재직", "입사일": "2021-05-17", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "제품본부", "소속부서": "플랫폼개발팀",
     "이름": "이서준", "직책": "Backend Engineer", "직급": "책임", "사번": "ENG002",
     "이메일": "seojun.lee@example.com", "재직상태": "재직", "입사일": "2022-09-05", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "제품본부", "소속부서": "프로덕트디자인팀",
     "이름": "윤아린", "직책": "Product Designer", "직급": "선임", "사번": "DSN001",
     "이메일": "arin.yoon@example.com", "재직상태": "재직", "입사일": "2023-11-06", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "사업본부", "소속부서": "세일즈팀",
     "이름": "문태오", "직책": "팀장", "직급": "책임", "사번": "SALES001",
     "이메일": "taeo.moon@example.com", "재직상태": "재직", "입사일": "2022-04-01", "퇴사일": ""},
    {"소속회사": TEMPLATE_COMPANY, "소속조직": "사업본부", "소속부서": "고객성공팀",
     "이름": "배수빈", "직책": "CS Manager", "직급": "매니저", "사번": "CS001",
     "이메일": "subin.bae@example.com", "재직상태": "재직", "입사일": "2024-03-04", "퇴사일": ""},
]

TEMPLATE_HIERARCHY_ROWS = [
    {"구분": "회사", "이름": TEMPLATE_COMPANY, "상위": "", "표시순서": 0},
    {"구분": "조직", "이름": "경영지원본부", "상위": TEMPLATE_COMPANY, "표시순서": 1},
    {"구분": "조직", "이름": "제품본부", "상위": TEMPLATE_COMPANY, "표시순서": 2},
    {"구분": "조직", "이름": "사업본부", "상위": TEMPLATE_COMPANY, "표시순서": 3},
    {"구분": "부서", "이름": "대표이사실", "상위": TEMPLATE_COMPANY, "표시순서": 0},
    {"구분": "부서", "이름": "People팀", "상위": "경영지원본부", "표시순서": 1},
    {"구분": "부서", "이름": "재무회계팀", "상위": "경영지원본부", "표시순서": 2},
    {"구분": "부서", "이름": "플랫폼개발팀", "상위": "제품본부", "표시순서": 1},
    {"구분": "부서", "이름": "프로덕트디자인팀", "상위": "제품본부", "표시순서": 2},
    {"구분": "부서", "이름": "세일즈팀", "상위": "사업본부", "표시순서": 1},
    {"구분": "부서", "이름": "고객성공팀", "상위": "사업본부", "표시순서": 2},
]

TEMPLATE_PEOPLE_SHEET = "명단"
TEMPLATE_HIERARCHY_SHEET = "위계"


# 헤더 서식(DESIGN.md accent 계열) 및 컬럼 폭 상수.
_HEADER_FILL = PatternFill("solid", fgColor="A0002A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_PEOPLE_WIDTHS = {
    "소속회사": 20, "소속조직": 16, "소속부서": 16, "이름": 12, "직책": 18,
    "직급": 10, "사번": 12, "이메일": 26, "재직상태": 10, "입사일": 13, "퇴사일": 13,
}
_HIERARCHY_WIDTHS = {"구분": 8, "이름": 22, "상위": 22, "표시순서": 10}


def _style_header_sheet(worksheet, widths: dict[str, int]) -> None:
    """헤더 행 서식·틀고정·컬럼 폭을 적용한다."""
    for col_index, cell in enumerate(worksheet[1], start=1):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        header = str(cell.value) if cell.value is not None else ""
        width = widths.get(header, 14)
        worksheet.column_dimensions[get_column_letter(col_index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 22


def _add_dropdown(worksheet, column_index: int, options: list[str], max_rows: int = 400) -> None:
    """지정 컬럼에 값 목록 데이터 유효성(드롭다운)을 건다."""
    formula = '"' + ",".join(options) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "목록에 있는 값만 입력할 수 있습니다."
    validation.errorTitle = "잘못된 값"
    validation.prompt = "목록에서 선택하세요: " + ", ".join(options)
    validation.promptTitle = "값 선택"
    worksheet.add_data_validation(validation)
    letter = get_column_letter(column_index)
    validation.add(f"{letter}2:{letter}{max_rows}")


def write_org_template(path: Path, *, with_samples: bool = True) -> Path:
    """표준 조직도 템플릿(명단+위계 2 시트) Excel을 생성한다.

    헤더 서식·틀고정·컬럼 폭·데이터 유효성(드롭다운)까지 적용해
    비개발자가 열자마자 채워 넣을 수 있는 양식으로 만든다.
    with_samples=False이면 헤더만 있는 빈 양식을 만든다.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    people_rows = TEMPLATE_PEOPLE_ROWS if with_samples else []
    hierarchy_rows = TEMPLATE_HIERARCHY_ROWS if with_samples else []
    people_columns = list(TEMPLATE_PEOPLE_ROWS[0].keys())
    hierarchy_columns = list(TEMPLATE_HIERARCHY_ROWS[0].keys())
    people_frame = pd.DataFrame(people_rows, columns=people_columns)
    hierarchy_frame = pd.DataFrame(hierarchy_rows, columns=hierarchy_columns)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        people_frame.to_excel(writer, sheet_name=TEMPLATE_PEOPLE_SHEET, index=False)
        hierarchy_frame.to_excel(writer, sheet_name=TEMPLATE_HIERARCHY_SHEET, index=False)

        people_ws = writer.sheets[TEMPLATE_PEOPLE_SHEET]
        hierarchy_ws = writer.sheets[TEMPLATE_HIERARCHY_SHEET]
        _style_header_sheet(people_ws, _PEOPLE_WIDTHS)
        _style_header_sheet(hierarchy_ws, _HIERARCHY_WIDTHS)

        # 명단!재직상태 · 위계!구분 드롭다운(데이터 유효성).
        if "재직상태" in people_columns:
            _add_dropdown(people_ws, people_columns.index("재직상태") + 1, ["재직", "휴직", "퇴사"])
        if "구분" in hierarchy_columns:
            _add_dropdown(hierarchy_ws, hierarchy_columns.index("구분") + 1, ["회사", "조직", "부서"])
    return path
